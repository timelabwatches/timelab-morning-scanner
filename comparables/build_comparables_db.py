#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
TIMELAB — build_comparables_db.py

Builds a hierarchical Catawiki comparables database from:
  - Q1 Excel (ground truth: brand, model, mechanism, supplier, real hammer)
  - Catawiki export (154 entregados/enviados; primary source for hammers)

Output: comparables_db_v2.json with:
  - hierarchical buckets (level 1 ref → level 5 brand fallback)
  - tier multipliers per brand (vintage/essentials/standard, learned)
  - chrono multipliers per brand
  - mechanism multipliers per brand

Usage:
  python build_comparables_db.py --q1 TL_Contabilidad_1T26.xlsx \
                                 --catawiki Pedidos_export.xlsx \
                                 --out comparables_db_v2.json

Dimensions captured per closed sale:
  brand, model_text, mechanism, is_chrono, auction_tier, hammer
"""

import argparse
import json
import re
import statistics
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

# Support both `python -m comparables.build_comparables_db ...` (package) and
# direct `python build_comparables_db.py ...` (script in the same folder).
try:
    from .enrichers import infer_model_family, infer_mechanism_from_refs
except ImportError:
    from enrichers import infer_model_family, infer_mechanism_from_refs


# --------- Config ----------

MIN_BUCKET_N = 4               # don't publish a bucket with less than 4 comps
TIER_MIN_N_FOR_MULT = 4        # need at least 4 sales in a tier within a brand to learn multiplier
DEFAULT_MULT = 1.0             # if signal too weak, neutral

NORM_BRAND_MAP = {
    # Lowercase synonyms → canonical
    "tissot": "Tissot", "seiko": "Seiko", "longines": "Longines",
    "omega": "Omega", "hamilton": "Hamilton", "zenith": "Zenith",
    "tag heuer": "TAG Heuer", "tagheuer": "TAG Heuer",
    "baume & mercier": "Baume & Mercier", "baume": "Baume & Mercier",
    "junghans": "Junghans", "certina": "Certina", "oris": "Oris",
    "rado": "Rado", "fortis": "Fortis", "cyma": "Cyma",
    "cauny": "Cauny", "citizen": "Citizen", "mido": "Mido",
    "mortima": "Mortima", "dogma": "Dogma",
    "favre-leuba": "Favre-Leuba", "favre leuba": "Favre-Leuba",
    "universal geneve": "Universal Genève", "universal genève": "Universal Genève",
    "maurice lacroix": "Maurice Lacroix",
    "tag-heuer": "TAG Heuer",
}

REF_PATTERNS = [
    r"\bT\d{3}\.\d{3}\b",                  # Tissot T120.417
    r"\bT\d{6}\b",                         # Tissot T127410
    r"\bT-?\d{3}-?\d{3}\b",                # Tissot T-048417
    r"\bSRP[A-Z]\d{2,3}[A-Z]?\d?\b",       # Seiko SRPK87, SRPF37J1
    r"\bSNX[A-Z]?\d{2,3}\b",               # Seiko SNXS73
    r"\bSNK\d{2,3}[A-Z]?\d?\b",            # Seiko SNK357K1
    r"\b\d{4}-\d{4}\b",                    # Seiko vintage 7005-8190, 6306-8001
    r"\b6[A-Z]?\d{2}-\d{4}\b",             # Seiko 6G34-00B0
    r"\b\d{3}\.\d{3}\.\d{2}\.\d{3}\.\d{2}\b",  # Tissot full ref
    r"\b[A-Z]\d{6}\b",                     # Hamilton H685820
]

CHRONO_TERMS = ["chrono", "chronograph", "cronograf", "chronograf"]

MECH_INFER_RULES = [
    (("automatic", "automatico", "automático", "self-winding", "self winding", "self windng"), "Automático"),
    (("quartz", "cuarzo"), "Cuarzo"),
    (("electronic", "electronico", "electrónico", "tuning fork", "f300", "esa 9162"), "Electrónico"),
    (("hand-wound", "hand wound", "handwound", "manual wind", "manual winding", "wind", "cuerda"), "Cuerda"),
]


# --------- Helpers ----------

def safe_str(v):
    return "" if v is None else str(v).strip()

def norm_brand(raw: str) -> str:
    if not raw:
        return ""
    k = raw.lower().strip()
    return NORM_BRAND_MAP.get(k, raw.strip())

def _looks_like_year_range(s: str) -> bool:
    """Catawiki object names contain '1970-1979' style year ranges; exclude these."""
    m = re.fullmatch(r"(\d{4})-(\d{4})", s)
    if not m:
        return False
    a, b = int(m.group(1)), int(m.group(2))
    return 1900 <= a <= 2050 and 1900 <= b <= 2050 and 0 < (b - a) <= 30

def extract_refs(text: str) -> list:
    refs = set()
    t = (text or "").upper().replace("–", "-")
    for pat in REF_PATTERNS:
        for m in re.finditer(pat, t, re.IGNORECASE):
            ref = m.group(0).upper().replace(" ", "")
            if _looks_like_year_range(ref):
                continue
            refs.add(ref)
    return sorted(refs)

def is_chrono(text: str) -> bool:
    t = (text or "").lower()
    return any(c in t for c in CHRONO_TERMS)

def infer_mech(text: str) -> str:
    t = (text or "").lower()
    for terms, mech in MECH_INFER_RULES:
        if any(term in t for term in terms):
            return mech
    return "?"

def infer_auction_tier(auction_name: str) -> str:
    if not auction_name:
        return "standard"
    a = auction_name.lower()
    if "premium" in a:
        return "premium"
    if "vintage" in a:
        return "vintage"
    if "esenciales" in a or "menos de 1500" in a:
        return "essentials"
    return "standard"

def percentiles(prices: list) -> dict:
    s = sorted(prices)
    n = len(s)
    if n == 0:
        return None
    return {
        "n": n,
        "min": round(s[0], 2),
        "p25": round(s[max(0, n // 4)], 2),
        "p50": round(s[n // 2], 2),
        "p75": round(s[min(n - 1, 3 * n // 4)], 2),
        "p90": round(s[min(n - 1, int(0.9 * n))], 2),
        "max": round(s[-1], 2),
        "mean": round(statistics.mean(s), 2),
        "stdev": round(statistics.stdev(s), 2) if n >= 2 else 0.0,
    }


# --------- Loaders ----------

def load_excel_catawiki_ops(path: Path, source_label: str) -> list:
    """Load Catawiki-channel sold ops from any TL Contabilidad Excel.
    Same schema for Q1, Q2, future quarters."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["OPERACIONES"]
    rows = list(ws.iter_rows(values_only=True))
    H = {h: i for i, h in enumerate(rows[0])}
    out = []
    for r in rows[1:]:
        if not r or not r[H["MARCA"]]:
            continue
        sale = float(r[H["PRECIO_VENTA_RELOJ"]] or 0)
        if sale <= 0:
            continue
        if safe_str(r[H["CANAL VENTA"]]) != "Catawiki":
            continue
        brand = norm_brand(safe_str(r[H["MARCA"]]))
        model = safe_str(r[H["MODELO"]])
        mech = safe_str(r[H["MECANISMO"]])
        family = infer_model_family(brand, f"{brand} {model}")
        out.append({
            "src": source_label,
            "brand": brand,
            "model_text": model,
            "model_family": family,
            "mechanism": mech,
            "is_chrono": is_chrono(model),
            "refs": extract_refs(model),
            "auction_tier": "?",
            "supplier_type": safe_str(r[H["TIPO PROVEEDOR"]]),
            "buy": float(r[H["COSTE TOTAL COMPRA"]] or 0),
            "hammer": sale,
            "fecha_venta": r[H["FECHA_VENTA"]],
        })
    wb.close()
    return out


def load_q1_catawiki_ops(path: Path) -> list:
    """Backward-compat alias."""
    return load_excel_catawiki_ops(path, source_label="Q1")


def load_catawiki_export(path: Path) -> list:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    HEADERS = rows[0]
    out = []
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        status = safe_str(r[12])
        if status not in ("Entregado", "Enviado por correo al comprador"):
            continue
        hammer = r[31]
        if not hammer:
            continue
        name = safe_str(r[0])
        parts = [p.strip() for p in name.split(" - ")]
        brand = norm_brand(parts[0] if parts else "")
        # auction info column 8
        auction = safe_str(r[8])
        refs = extract_refs(name)
        # Mechanism: try the listing-name terms first, then upgrade via ref→mech
        mech_from_text = infer_mech(name)
        mech_from_refs = infer_mechanism_from_refs(refs, brand_hint=brand) if refs else None
        if mech_from_text != "?":
            mechanism = mech_from_text
            mech_source = "name_text"
        elif mech_from_refs:
            mechanism = mech_from_refs
            mech_source = "ref_lookup"
        else:
            mechanism = "?"
            mech_source = "unknown"
        family = infer_model_family(brand, name)
        out.append({
            "src": "CW",
            "brand": brand,
            "object_name": name,
            "model_text": " ".join(parts[1:3]) if len(parts) >= 3 else "",
            "model_family": family,
            "mechanism": mechanism,
            "mech_source_initial": mech_source,
            "is_chrono": is_chrono(name),
            "refs": refs,
            "auction_tier": infer_auction_tier(auction),
            "buy": None,
            "hammer": float(hammer),
            "fecha_venta": r[10],
            "country": safe_str(r[28]),
        })
    wb.close()
    return out


# --------- Merge & dedupe ----------

def merge_sources(excel_truth: list, catawiki: list) -> list:
    """
    Merge multiple Excel ground-truth ops with the Catawiki export.
    Catawiki is the canonical price source. Excel rows enrich with mechanism
    ground truth when they match (same brand + same hammer).

    excel_truth: list of dicts (Q1 + T2 + ... combined)
    catawiki:    list from load_catawiki_export()

    Avoids double-counting by tracking matched indices.
    """
    enriched = []
    matched_idx = set()

    for cw in catawiki:
        match = None
        for idx, q in enumerate(excel_truth):
            if idx in matched_idx:
                continue
            if q["brand"] != cw["brand"]:
                continue
            if abs(q["hammer"] - cw["hammer"]) > 0.5:
                continue
            match = (idx, q)
            break
        if match:
            matched_idx.add(match[0])
            cw_copy = dict(cw)
            # ground truth wins for mechanism
            if match[1]["mechanism"]:
                cw_copy["mechanism"] = match[1]["mechanism"]
                cw_copy["mech_source"] = f"{match[1]['src']}_truth"
            cw_copy["model_text_truth"] = match[1]["model_text"]
            cw_copy["supplier_type"] = match[1]["supplier_type"]
            cw_copy["buy"] = match[1]["buy"]
            cw_copy["matched"] = True
            # NOTE: model_family is NOT inherited from truth.
            # Family is a property of how the watch was LISTED on Catawiki,
            # not of how it appears in our internal spreadsheet (which can use
            # different naming conventions, model variants, or typos).
            enriched.append(cw_copy)
        else:
            cw_copy = dict(cw)
            cw_copy["mech_source"] = "inferred" if cw_copy["mechanism"] != "?" else "unknown"
            cw_copy["matched"] = False
            enriched.append(cw_copy)

    # Excel-only ops (no matching CW row — earlier dates outside export, edge cases)
    for idx, q in enumerate(excel_truth):
        if idx in matched_idx:
            continue
        q_copy = dict(q)
        q_copy["mech_source"] = f"{q['src']}_truth"
        q_copy["matched"] = False
        if q_copy.get("auction_tier") == "?":
            q_copy["auction_tier"] = "standard"
        enriched.append(q_copy)

    return enriched


# --------- Bucket builder ----------

def build_buckets(comps: list) -> dict:
    """
    Hierarchy levels:
      L1  ref:{ref}                                    -- exact reference
      L2  brand_family_mech_chrono:{B}:{F}:{X}:{C}     -- brand × family × mech × chrono
      L3  brand_mech_chrono:{B}:{X}:{C}                -- brand × mech × chrono
      L4  brand_mech:{B}:{X}                           -- brand × mech
      L5  brand:{B}                                    -- brand fallback
    """
    buckets = {}

    by_ref = defaultdict(list)
    by_bfmc = defaultdict(list)
    by_bmc = defaultdict(list)
    by_bm = defaultdict(list)
    by_b = defaultdict(list)

    for c in comps:
        h = c["hammer"]
        b = c["brand"]
        if not b:
            continue
        m = c.get("mechanism", "?")
        chrono = "Y" if c.get("is_chrono") else "N"
        family = c.get("model_family")
        for ref in c.get("refs", []):
            by_ref[ref].append(h)
        if family and m != "?":
            by_bfmc[(b, family, m, chrono)].append(h)
        by_bmc[(b, m, chrono)].append(h)
        by_bm[(b, m)].append(h)
        by_b[b].append(h)

    for ref, prices in by_ref.items():
        if len(prices) >= MIN_BUCKET_N:
            buckets[f"L1:ref:{ref}"] = {
                "level": 1, "key_type": "reference", "key": ref,
                **percentiles(prices),
            }

    for (b, f, m, c), prices in by_bfmc.items():
        if len(prices) >= MIN_BUCKET_N:
            buckets[f"L2:brand_family_mech_chrono:{b}:{f}:{m}:{c}"] = {
                "level": 2, "key_type": "brand_family_mech_chrono",
                "brand": b, "family": f, "mechanism": m, "is_chrono": (c == "Y"),
                **percentiles(prices),
            }

    for (b, m, c), prices in by_bmc.items():
        if len(prices) >= MIN_BUCKET_N and m != "?":
            buckets[f"L3:brand_mech_chrono:{b}:{m}:{c}"] = {
                "level": 3, "key_type": "brand_mech_chrono",
                "brand": b, "mechanism": m, "is_chrono": (c == "Y"),
                **percentiles(prices),
            }

    for (b, m), prices in by_bm.items():
        if len(prices) >= MIN_BUCKET_N and m != "?":
            buckets[f"L4:brand_mech:{b}:{m}"] = {
                "level": 4, "key_type": "brand_mech",
                "brand": b, "mechanism": m,
                **percentiles(prices),
            }

    for b, prices in by_b.items():
        if len(prices) >= MIN_BUCKET_N:
            buckets[f"L5:brand:{b}"] = {
                "level": 5, "key_type": "brand", "brand": b,
                **percentiles(prices),
            }

    return buckets


# --------- Multiplier learners ----------

def learn_tier_multipliers(comps: list) -> dict:
    """
    For each brand, compute median hammer per tier, then express as multiplier
    over the brand's standard-tier median.
    Only learn if base (standard) and tier each have >= TIER_MIN_N_FOR_MULT.
    """
    by_brand_tier = defaultdict(lambda: defaultdict(list))
    for c in comps:
        if c.get("auction_tier") in (None, "?"):
            continue
        by_brand_tier[c["brand"]][c["auction_tier"]].append(c["hammer"])

    out = {}
    for brand, tiers in by_brand_tier.items():
        std = tiers.get("standard", [])
        if len(std) < TIER_MIN_N_FOR_MULT:
            continue
        std_p50 = statistics.median(std)
        out[brand] = {"standard": 1.0, "_std_p50": round(std_p50, 2), "_std_n": len(std)}
        for tier in ("vintage", "essentials", "premium"):
            vals = tiers.get(tier, [])
            if len(vals) >= TIER_MIN_N_FOR_MULT:
                out[brand][tier] = round(statistics.median(vals) / std_p50, 3)
                out[brand][f"_{tier}_n"] = len(vals)
    return out


def learn_chrono_multipliers(comps: list) -> dict:
    """
    For each (brand, mech), compute median hammer for chrono vs non-chrono.
    Multiplier = chrono_p50 / nonchrono_p50, when both have >= TIER_MIN_N_FOR_MULT.
    """
    by_bmc = defaultdict(lambda: defaultdict(list))
    for c in comps:
        m = c.get("mechanism", "?")
        if m == "?":
            continue
        flag = "chrono" if c.get("is_chrono") else "non"
        by_bmc[(c["brand"], m)][flag].append(c["hammer"])

    out = {}
    for (brand, mech), groups in by_bmc.items():
        chrono_p = groups.get("chrono", [])
        non_p = groups.get("non", [])
        if len(chrono_p) >= TIER_MIN_N_FOR_MULT and len(non_p) >= TIER_MIN_N_FOR_MULT:
            mult = statistics.median(chrono_p) / statistics.median(non_p)
            out[f"{brand}:{mech}"] = {
                "multiplier": round(mult, 3),
                "chrono_n": len(chrono_p),
                "non_n": len(non_p),
                "chrono_p50": round(statistics.median(chrono_p), 2),
                "non_p50": round(statistics.median(non_p), 2),
            }
    return out


def learn_mech_multipliers(comps: list) -> dict:
    """
    For each brand, compute median hammer per mechanism, expressed as
    multiplier over Cuarzo (most common cross-brand reference).
    Only emitted when brand has >= TIER_MIN_N_FOR_MULT in each mech.
    """
    by_brand_mech = defaultdict(lambda: defaultdict(list))
    for c in comps:
        m = c.get("mechanism", "?")
        if m == "?":
            continue
        by_brand_mech[c["brand"]][m].append(c["hammer"])

    out = {}
    for brand, mechs in by_brand_mech.items():
        cuarzo = mechs.get("Cuarzo", [])
        if len(cuarzo) < TIER_MIN_N_FOR_MULT:
            continue
        cuarzo_p50 = statistics.median(cuarzo)
        if cuarzo_p50 <= 0:
            continue
        rec = {"Cuarzo": 1.0, "_cuarzo_p50": round(cuarzo_p50, 2)}
        for m in ("Automático", "Cuerda", "Electrónico"):
            vals = mechs.get(m, [])
            if len(vals) >= TIER_MIN_N_FOR_MULT:
                rec[m] = round(statistics.median(vals) / cuarzo_p50, 3)
                rec[f"_{m}_n"] = len(vals)
        out[brand] = rec
    return out


def learn_family_mech_defaults(comps: list, min_n: int = 3) -> dict:
    """
    Per (brand, family), if all comps share the SAME mechanism AND n >= min_n,
    publish that as a safe default. Returns:
        { "Brand:family": {"mechanism": "Cuarzo", "n": 5, "source": "data_driven"} }

    Used at runtime by the enricher when the analyzer fails to extract movement
    from listing text and ref-lookup also fails.
    """
    by_bf = defaultdict(list)
    for c in comps:
        family = c.get("model_family")
        mech = c.get("mechanism")
        if not family or mech in (None, "", "?"):
            continue
        by_bf[(c["brand"], family)].append(mech)

    out = {}
    for (brand, family), mechs in by_bf.items():
        if len(mechs) < min_n:
            continue
        unique = set(mechs)
        if len(unique) == 1:
            out[f"{brand}:{family}"] = {
                "mechanism": mechs[0],
                "n": len(mechs),
                "source": "data_driven",
            }
    return out


# --------- Main ----------

def build_db(excel_paths: list, catawiki_path: Path, out_path: Path) -> dict:
    """
    excel_paths: list of (path, label) tuples e.g. [("Q1.xlsx","Q1"), ("Q2.xlsx","Q2")]
    """
    excel_truth = []
    for p, label in excel_paths:
        ops = load_excel_catawiki_ops(Path(p), source_label=label)
        print(f"[load] {label} Catawiki ops: {len(ops)}")
        excel_truth.extend(ops)

    cw = load_catawiki_export(Path(catawiki_path))
    print(f"[load] Catawiki entregados/enviados: {len(cw)}")

    comps = merge_sources(excel_truth, cw)
    n_matched = sum(1 for c in comps if c.get("matched"))
    print(f"[merge] unified comps: {len(comps)} (matched Excel↔CW: {n_matched})")

    buckets = build_buckets(comps)
    tier_mults = learn_tier_multipliers(comps)
    chrono_mults = learn_chrono_multipliers(comps)
    mech_mults = learn_mech_multipliers(comps)
    family_mech_defaults = learn_family_mech_defaults(comps)

    db = {
        "version": "v2-catawiki-only",
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "scope": "Catawiki hammer prices only",
        "sources": {
            "excel": [str(p) for p, _ in excel_paths],
            "catawiki_export": str(catawiki_path),
        },
        "n_comps_total": len(comps),
        "n_truth_mech": sum(1 for c in comps if "truth" in (c.get("mech_source") or "")),
        "n_inferred":   sum(1 for c in comps if c.get("mech_source") == "inferred"),
        "n_unknown_mech": sum(1 for c in comps if c.get("mech_source") == "unknown"),
        "min_bucket_n": MIN_BUCKET_N,
        "tier_multipliers_by_brand": tier_mults,
        "chrono_multipliers_by_brand_mech": chrono_mults,
        "mech_multipliers_by_brand": mech_mults,
        "family_mech_defaults": family_mech_defaults,
        "buckets": buckets,
    }

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(db, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[write] {out_path}  ({len(buckets)} buckets)")
    return db


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--excel", action="append", default=[],
                   help="Excel source as PATH:LABEL (repeat for multiple). "
                        "Example: --excel TL_1T26.xlsx:Q1 --excel TL_2T26.xlsx:Q2")
    # Backward-compat single-Q1 flag
    p.add_argument("--q1", help="(legacy) Q1 Excel path; same as --excel <path>:Q1")
    p.add_argument("--catawiki", required=True, help="Catawiki export Excel path")
    p.add_argument("--out", default="comparables_db_v2.json")
    args = p.parse_args()

    excel_paths = []
    for spec in args.excel:
        if ":" in spec:
            pth, lab = spec.rsplit(":", 1)
        else:
            pth, lab = spec, Path(spec).stem
        excel_paths.append((pth, lab))
    if args.q1:
        excel_paths.insert(0, (args.q1, "Q1"))

    if not excel_paths:
        raise SystemExit("Need at least one --excel or --q1 source.")

    build_db(excel_paths, Path(args.catawiki), Path(args.out))


if __name__ == "__main__":
    main()
